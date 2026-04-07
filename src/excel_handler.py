import enum

import openpyxl
import pandas as pd

from src import config

class DataSetType(enum.Enum):
    RAW = "raw"
    TEMP = "temp"
    PROCESSED = "processed"

class ExcelHandler:    
    cfg: config.Excel
    raw_workbook: openpyxl.workbook.workbook.Workbook
    temp_workbook: openpyxl.workbook.workbook.Workbook
    processed_workbook: openpyxl.workbook.workbook.Workbook
    df: pd.DataFrame
    
    def __init__(self, cfg: config.Excel):
        self.cfg = cfg
        self.raw_workbook = openpyxl.load_workbook(self.cfg.raw_path)
        self.temp_workbook = openpyxl.load_workbook(self.cfg.temp_path)
        self.processed_workbook = openpyxl.load_workbook(self.cfg.processed_path)
        
    def save(self, type: DataSetType, df: pd.DataFrame) -> None:
        match type:
            case DataSetType.RAW:
                df.to_excel(
                    self.cfg.raw_path,
                    sheet_name=self.cfg.raw_sheet_name,
                    index=False,
                    engine='openpyxl',
                )
            case DataSetType.TEMP:
                df.to_excel(
                    self.cfg.temp_path,
                    sheet_name=self.cfg.temp_sheet_name,
                    index=False,
                    engine='openpyxl',
                )
            case DataSetType.PROCESSED:
                df.to_excel(
                    self.cfg.processed_path,
                    sheet_name=self.cfg.processed_sheet_name,
                    index=False,
                    engine='openpyxl',
                )
            case _:
                raise ValueError(f"Unsupported dataset type: {type}")
                
    def get_df(self, type: DataSetType) -> pd.DataFrame:
        match type:
            case DataSetType.RAW:
                return pd.read_excel(self.cfg.raw_path, sheet_name=self.cfg.raw_sheet_name)
            case DataSetType.TEMP:
                return pd.read_excel(self.cfg.temp_path, sheet_name=self.cfg.temp_sheet_name)
            case DataSetType.PROCESSED:
                return pd.read_excel(self.cfg.processed_path, sheet_name=self.cfg.processed_sheet_name)
            case _:
                raise ValueError(f"Unsupported dataset type: {type}")

    def clear(self, type: DataSetType) -> None:
        match type:
            case DataSetType.RAW:
                self.raw_workbook[self.cfg.raw_sheet_name].delete_rows(1, self.raw_workbook[self.cfg.raw_sheet_name].max_row)
                self.raw_workbook.save(self.cfg.raw_path)
            case DataSetType.TEMP:
                self.temp_workbook[self.cfg.temp_sheet_name].delete_rows(1, self.temp_workbook[self.cfg.temp_sheet_name].max_row)
                self.temp_workbook.save(self.cfg.temp_path)
            case DataSetType.PROCESSED:
                self.processed_workbook[self.cfg.processed_sheet_name].delete_rows(1, self.processed_workbook[self.cfg.processed_sheet_name].max_row)
                self.processed_workbook.save(self.cfg.processed_path)
            case _:
                raise ValueError(f"Unsupported dataset type: {type}")
                  
    def set_raw_path(self, raw_path: str) -> None:
        self.raw_workbook = openpyxl.load_workbook(raw_path)
        self.cfg.raw_path = raw_path